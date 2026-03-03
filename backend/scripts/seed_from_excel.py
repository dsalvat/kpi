"""
Seed database from KPIs_OKRs_IT_Sistemes.xlsx

Reads:
  '📈 KPIs Serveis'    -> kpi_definitions + kpi_values (year 2025)
  '🚀 KPIs Projectes'  -> projects + project KPI defs/values
  '💰 Valor Negoci'    -> value_items Blue/Green Money
  '🎯 OKRs'            -> okr_objectives + okr_key_results + okr_quarterly_data

Usage:
  docker compose exec backend python -m scripts.seed_from_excel /data/KPIs_OKRs_IT_Sistemes.xlsx

Note: Uses synchronous SQLAlchemy since this runs as a one-time script.
"""

import sys
from datetime import date, datetime, timezone
from decimal import Decimal, InvalidOperation
from pathlib import Path

import openpyxl
from sqlalchemy import create_engine, text
from sqlalchemy.orm import Session

# Fix import path
sys.path.insert(0, str(Path(__file__).resolve().parent.parent))

from app.config import settings
from app.database import Base
from app.models.kpi import KPIDefinition, KPIValue
from app.models.okr import OKRKeyResult, OKRObjective, OKRQuarterlyData
from app.models.project import Project
from app.models.value import ValueItem

# Build sync URL from async URL
_db_url = settings.DATABASE_URL
if "sqlite" in _db_url:
    SYNC_URL = _db_url.replace("+aiosqlite", "").replace("aiosqlite", "sqlite")
else:
    SYNC_URL = _db_url.replace("+asyncpg", "+psycopg2").replace("asyncpg", "psycopg2")

YEAR = 2025

# Explicit lower_better KPI codes (all others are higher_better)
# SRV-03/04/05: resolution times (h), SRV-07/08: incident counts, SRV-11: cost per ticket (€)
# PRJ-M02: budget deviation (lower = better), PRJ-M04: hours vs planned (closer to 100%)
LOWER_BETTER_CODES = {
    "SRV-03", "SRV-04", "SRV-05", "SRV-07", "SRV-08", "SRV-11",
    "PRJ-M02", "PRJ-M04",
}

# Map Excel status emojis to model status values
STATUS_MAP = {
    "Tancat": "completed",
    "En curs": "active",
    "En risc": "at_risk",
    "Bloquejat": "blocked",
}


def _dec(value) -> Decimal | None:
    """Convert Excel cell value to Decimal, returning None for empty/invalid."""
    if value is None or value == "" or value == "—" or value == "-":
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _parse_date(value) -> date | None:
    """Convert Excel date cell to Python date."""
    if value is None or value == "-" or value == "—":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    # Try parsing string dates like "01/01/2025"
    if isinstance(value, str):
        for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def _parse_status(value: str | None) -> str:
    """Parse project status from Excel (e.g. '✅ Tancat' -> 'completed')."""
    if not value:
        return "active"
    for key, status in STATUS_MAP.items():
        if key in str(value):
            return status
    return "active"


def _direction(code: str, unit: str) -> str:
    """Determine KPI direction based on code."""
    if code in LOWER_BETTER_CODES:
        return "lower_better"
    return "higher_better"


def _source_for_code(code: str) -> str:
    """Map KPI code to data source."""
    source_map = {
        "SRV-01": "meraki",
        "SRV-02": "aruba",
        "SRV-03": "itsm",
        "SRV-04": "itsm",
        "SRV-05": "itsm",
        "SRV-06": "itsm",
        "SRV-07": "itsm",
        "SRV-08": "itsm",
        "SRV-09": "itsm",
        "SRV-10": "itsm",
        "SRV-11": "itsm",
        "SRV-12": "itsm",
    }
    return source_map.get(code, "manual")


def seed_kpis_serveis(ws, session: Session) -> dict[str, KPIDefinition]:
    """Parse '📈 KPIs Serveis' sheet. Returns code->KPIDefinition map."""
    print("\n--- KPIs Serveis ---")
    kpi_map: dict[str, KPIDefinition] = {}

    # Headers at row 3, data rows 4-15 (SRV-01 to SRV-12)
    for row in range(4, 16):
        code = ws.cell(row, 1).value
        if not code or not str(code).startswith("SRV"):
            continue

        name = ws.cell(row, 2).value
        category = ws.cell(row, 3).value or "serveis"
        unit = ws.cell(row, 4).value or "%"
        target = _dec(ws.cell(row, 5).value)

        direction = _direction(code, unit)
        source = _source_for_code(code)

        kpi = KPIDefinition(
            code=code,
            name=name,
            category=category.lower() if category else "serveis",
            unit=unit,
            target=target,
            direction=direction,
            source=source,
            active=True,
        )
        session.add(kpi)
        session.flush()
        kpi_map[code] = kpi

        # Monthly values in columns 7 (Gen) to 18 (Des)
        values_count = 0
        for month, col in enumerate(range(7, 19), 1):
            val = _dec(ws.cell(row, col).value)
            if val is not None:
                session.add(KPIValue(
                    kpi_id=kpi.id,
                    year=YEAR,
                    month=month,
                    value=val,
                    collected_at=datetime(YEAR, month, 15, tzinfo=timezone.utc),
                    collection_method="manual",
                ))
                values_count += 1

        print(f"  {code}: {name} (target={target}, {values_count} monthly values)")

    print(f"  Total: {len(kpi_map)} KPI definitions")
    return kpi_map


def seed_projects(ws, session: Session) -> dict[str, KPIDefinition]:
    """Parse '🚀 KPIs Projectes' sheet. Returns code->KPIDefinition map for project KPIs."""
    print("\n--- Projects ---")
    kpi_map: dict[str, KPIDefinition] = {}

    # Section A: Individual projects, rows 5-12 (PRJ-01 to PRJ-08)
    project_count = 0
    for row in range(5, 13):
        code = ws.cell(row, 1).value
        if not code or not str(code).startswith("PRJ-"):
            continue
        # Skip PRJ-M codes (they're in section B)
        if "M" in str(code):
            continue

        name = ws.cell(row, 2).value
        sponsor = ws.cell(row, 3).value
        status_raw = ws.cell(row, 4).value
        start_date = _parse_date(ws.cell(row, 5).value)
        planned_end = _parse_date(ws.cell(row, 6).value)
        actual_end = _parse_date(ws.cell(row, 7).value)
        completion = ws.cell(row, 8).value
        budget = _dec(ws.cell(row, 9).value)
        actual_cost = _dec(ws.cell(row, 10).value)
        satisfaction = _dec(ws.cell(row, 15).value)

        # Convert completion from decimal (0.75) to percentage (75)
        completion_pct = None
        if completion is not None:
            try:
                pct = float(completion)
                if pct <= 1:
                    pct = pct * 100
                completion_pct = int(pct)
            except (ValueError, TypeError):
                pass

        status = _parse_status(status_raw)

        project = Project(
            code=code,
            name=name,
            sponsor=sponsor,
            status=status,
            start_date=start_date,
            planned_end_date=planned_end,
            actual_end_date=actual_end,
            budget_eur=budget,
            actual_cost_eur=actual_cost,
            completion_pct=completion_pct,
            sponsor_satisfaction=satisfaction,
        )
        session.add(project)
        project_count += 1
        print(f"  {code}: {name} ({status}, {completion_pct}%)")

    print(f"  Total: {project_count} projects")

    # Section B: Monthly delivery KPIs, rows 17-24 (PRJ-M01 to PRJ-M08)
    print("\n--- Project Monthly KPIs ---")
    for row in range(17, 25):
        code = ws.cell(row, 1).value
        if not code or not str(code).startswith("PRJ-M"):
            continue

        name = ws.cell(row, 2).value
        category = ws.cell(row, 3).value or "projectes"
        unit = ws.cell(row, 4).value or "%"
        target = _dec(ws.cell(row, 5).value)

        direction = _direction(code, unit)

        kpi = KPIDefinition(
            code=code,
            name=name,
            category=category.lower() if category else "projectes",
            unit=unit,
            target=target,
            direction=direction,
            source="manual",
            active=True,
        )
        session.add(kpi)
        session.flush()
        kpi_map[code] = kpi

        # Monthly values in columns 7 (Gen) to 18 (Des)
        values_count = 0
        for month, col in enumerate(range(7, 19), 1):
            val = _dec(ws.cell(row, col).value)
            if val is not None:
                session.add(KPIValue(
                    kpi_id=kpi.id,
                    year=YEAR,
                    month=month,
                    value=val,
                    collected_at=datetime(YEAR, month, 15, tzinfo=timezone.utc),
                    collection_method="manual",
                ))
                values_count += 1

        print(f"  {code}: {name} (target={target}, {values_count} monthly values)")

    print(f"  Total: {len(kpi_map)} project KPI definitions")
    return kpi_map


def seed_value_items(ws, session: Session):
    """Parse '💰 Valor Negoci' sheet."""
    print("\n--- Blue Money ---")

    # Blue Money: rows 6-13 (BM-01 to BM-08)
    blue_count = 0
    for row in range(6, 14):
        code = ws.cell(row, 1).value
        if not code or not str(code).startswith("BM-"):
            continue

        initiative = ws.cell(row, 2).value
        hours_saved = _dec(ws.cell(row, 6).value)
        hourly_rate = _dec(ws.cell(row, 7).value)
        op_savings = _dec(ws.cell(row, 8).value)
        direct_savings = _dec(ws.cell(row, 9).value)

        item = ValueItem(
            type="blue_money",
            code=code,
            initiative=initiative or "",
            hours_saved_annual=hours_saved,
            hourly_rate_eur=hourly_rate,
            operational_savings_eur=op_savings,
            direct_savings_eur=direct_savings,
            year=YEAR,
        )
        session.add(item)
        blue_count += 1
        total = ws.cell(row, 10).value
        print(f"  {code}: {initiative} (total={total})")

    print(f"  Total: {blue_count} Blue Money items")

    # Green Money: rows 18-22 (GM-01 to GM-05)
    print("\n--- Green Money ---")
    green_count = 0
    for row in range(18, 23):
        code = ws.cell(row, 1).value
        if not code or not str(code).startswith("GM-"):
            continue

        initiative = ws.cell(row, 2).value
        revenues = _dec(ws.cell(row, 6).value)
        commercial_savings = _dec(ws.cell(row, 7).value)
        it_attribution = _dec(ws.cell(row, 9).value)

        item = ValueItem(
            type="green_money",
            code=code,
            initiative=initiative or "",
            revenues_enabled_eur=revenues,
            commercial_savings_eur=commercial_savings,
            it_attribution_pct=it_attribution,
            year=YEAR,
        )
        session.add(item)
        green_count += 1
        total = ws.cell(row, 10).value
        print(f"  {code}: {initiative} (total={total})")

    print(f"  Total: {green_count} Green Money items")


def seed_okrs(ws, session: Session, kpi_map: dict[str, KPIDefinition]):
    """Parse '🎯 OKRs' sheet."""
    print("\n--- OKRs ---")

    # Objective definitions with their row ranges and owners
    objectives_data = [
        {
            "code": "O1",
            "title": "Garantir l'excel·lència operativa i la fiabilitat dels sistemes crítics",
            "owner": "Dir. IT",
            "kr_rows": range(7, 11),  # O1.KR1 to O1.KR4
        },
        {
            "code": "O2",
            "title": "Generar ≥ 90.000€ en Blue Money (estalvis i eficiència operativa)",
            "owner": "Dir. IT",
            "kr_rows": range(15, 19),  # O2.KR1 to O2.KR4
        },
        {
            "code": "O3",
            "title": "Habilitar ≥ 400.000€ en nous ingressos gràcies a la tecnologia (Green Money)",
            "owner": "Dir. IT + Dir. Comercial",
            "kr_rows": range(23, 27),  # O3.KR1 to O3.KR4
        },
        {
            "code": "O4",
            "title": "Lliurar la cartera de projectes IT amb qualitat, impacte i dins de pressupost",
            "owner": "Dir. IT",
            "kr_rows": range(31, 35),  # O4.KR1 to O4.KR4
        },
        {
            "code": "O5",
            "title": "Enfortir la maduresa IT i les capacitats tècniques de l'equip",
            "owner": "Dir. IT",
            "kr_rows": range(39, 43),  # O5.KR1 to O5.KR4
        },
    ]

    for idx, obj_data in enumerate(objectives_data, 1):
        obj = OKRObjective(
            year=YEAR,
            code=obj_data["code"],
            title=obj_data["title"],
            owner=obj_data["owner"],
            order_idx=idx,
        )
        session.add(obj)
        session.flush()
        print(f"\n  {obj_data['code']}: {obj_data['title'][:60]}...")

        for row in obj_data["kr_rows"]:
            kr_code = ws.cell(row, 1).value
            if not kr_code or "KR" not in str(kr_code):
                continue

            title = ws.cell(row, 2).value or ""
            unit = ws.cell(row, 5).value or ""
            baseline = _dec(ws.cell(row, 6).value)
            annual_target = _dec(ws.cell(row, 7).value)

            if annual_target is None:
                continue

            # Parse confidence and KPI reference from column 21
            conf_cell = ws.cell(row, 21).value or ""
            confidence = None
            kpi_ref = None
            if conf_cell:
                lines = str(conf_cell).split("\n")
                if lines:
                    conf_text = lines[0].strip()
                    if "Alta" in conf_text:
                        confidence = "Alta"
                    elif "Mitjana" in conf_text or "Mittjana" in conf_text:
                        confidence = "Mitjana"
                    elif "Baixa" in conf_text:
                        confidence = "Baixa"
                if len(lines) > 1:
                    kpi_ref = lines[1].strip()

            # Determine direction from title keywords and unit
            direction = "higher_better"
            lower_keywords = ["reduir", "reduïr", "≤", "< ", "cost"]
            title_lower = title.lower()
            if unit == "h" or any(kw in title_lower for kw in lower_keywords):
                direction = "lower_better"

            # Link to KPI definition if reference matches
            kpi_id = None
            kpi_aggregation = None
            if kpi_ref:
                # Handle references like "SRV-01" or "BM-01/05" or "PRJ-M01"
                ref_code = kpi_ref.split("/")[0].strip()
                if ref_code in kpi_map:
                    kpi_id = kpi_map[ref_code].id
                    kpi_aggregation = "avg_quarter"

            kr = OKRKeyResult(
                objective_id=obj.id,
                code=kr_code,
                title=title,
                unit=unit,
                baseline=baseline,
                annual_target=annual_target,
                direction=direction,
                kpi_id=kpi_id,
                kpi_aggregation=kpi_aggregation,
                confidence=confidence,
            )
            session.add(kr)
            session.flush()

            # Quarterly data: Q1 (cols 8-10), Q2 (cols 11-13), Q3 (cols 14-16), Q4 (cols 17-19)
            quarters_info = []
            for q, (t_col, a_col) in enumerate([(8, 9), (11, 12), (14, 15), (17, 18)], 1):
                q_target = _dec(ws.cell(row, t_col).value)
                q_actual = _dec(ws.cell(row, a_col).value)

                if q_target is not None or q_actual is not None:
                    session.add(OKRQuarterlyData(
                        key_result_id=kr.id,
                        year=YEAR,
                        quarter=q,
                        target=q_target,
                        actual=q_actual,
                        is_manual=(q_actual is not None),
                    ))
                    quarters_info.append(f"Q{q}({q_target}/{q_actual})")

            kpi_label = f" -> {kpi_ref}" if kpi_ref else ""
            print(f"    {kr_code}: {title[:50]}... [{', '.join(quarters_info)}]{kpi_label}")


def clear_all_data(session: Session):
    """Delete all existing seed data (in correct FK order)."""
    print("Clearing existing data...")
    tables = [
        "okr_quarterly_data", "okr_key_results", "okr_objectives",
        "value_items", "kpi_values", "projects", "kpi_definitions", "n8n_sync_logs",
    ]
    for table in tables:
        try:
            session.execute(text(f"DELETE FROM {table}"))
        except Exception:
            pass  # Table may not exist yet
    session.commit()
    print("  Done.")


def seed_from_excel(excel_path: str, session: Session):
    """Parse Excel file and seed database with real data."""
    wb = openpyxl.load_workbook(excel_path, data_only=True)
    print(f"Excel loaded: {excel_path}")
    print(f"Sheets: {wb.sheetnames}")

    # Clear existing data
    clear_all_data(session)

    # 1. KPIs Serveis -> kpi_definitions + kpi_values
    ws_serveis = wb[wb.sheetnames[1]]  # '📈 KPIs Serveis'
    kpi_map = seed_kpis_serveis(ws_serveis, session)

    # 2. Projects + Project monthly KPIs
    ws_projectes = wb[wb.sheetnames[2]]  # '🚀 KPIs Projectes'
    prj_kpi_map = seed_projects(ws_projectes, session)

    # Merge KPI maps for OKR linking
    all_kpi_map = {**kpi_map, **prj_kpi_map}

    # 3. Value items (Blue Money + Green Money)
    ws_valor = wb[wb.sheetnames[3]]  # '💰 Valor Negoci'
    seed_value_items(ws_valor, session)

    # 4. OKRs
    ws_okrs = wb[wb.sheetnames[5]]  # '🎯 OKRs'
    seed_okrs(ws_okrs, session, all_kpi_map)

    session.commit()
    print("\n=== Seed completed successfully! ===")


def create_seed_engine():
    return create_engine(SYNC_URL)


def main():
    engine = create_seed_engine()

    # Create all tables (idempotent - safe if tables already exist)
    Base.metadata.create_all(engine)
    print("Database tables ensured.")

    with Session(engine) as session:
        if len(sys.argv) > 1 and Path(sys.argv[1]).exists():
            seed_from_excel(sys.argv[1], session)
        else:
            print("No Excel file provided or file not found.")
            print("Usage: python -m scripts.seed_from_excel <path_to_excel>")
            sys.exit(1)


if __name__ == "__main__":
    main()
